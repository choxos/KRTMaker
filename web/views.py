import os
import json
import uuid
import time
import tempfile
from datetime import datetime, timedelta

from django.shortcuts import render, get_object_or_404, redirect
from django.http import JsonResponse, HttpResponse, Http404
from django.views.decorators.csrf import csrf_exempt
from django.views.decorators.http import require_http_methods
from django.contrib import messages
from django.core.files.storage import default_storage
from django.utils.decorators import method_decorator
from django.views.generic import FormView, TemplateView
from django.db.models import Count, Avg, Sum, Q
from django.db import models
from django.utils import timezone

from .forms import KRTMakerForm, FeedbackForm
from .models import KRTSession, ProcessedFile, KRTExport, SystemMetrics, Article, XMLFile, AdminKRT

# Import the KRT maker functionality (using project root path)
from django.conf import settings
import sys
import os

# Add project root to path only once during module import
project_root = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))
if project_root not in sys.path:
    sys.path.insert(0, project_root)

from builder import build_from_xml_path, BuildOptions
from validation import validate_xml_file, validate_api_config, ValidationError as KRTValidationError
from biorxiv_fetcher import BioRxivFetcher
from krt_validation import validate_krt_completeness, get_krt_quality_score, suggest_krt_improvements


class HomeView(TemplateView):
    """Beautiful homepage with features and getting started info"""
    template_name = 'web/home.html'
    
    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        
        # Get some basic stats for the homepage
        total_sessions = KRTSession.objects.count()
        successful_sessions = KRTSession.objects.filter(status='completed').count()
        total_resources = KRTSession.objects.aggregate(Sum('resources_found'))['resources_found__sum'] or 0
        
        context.update({
            'total_sessions': total_sessions,
            'successful_sessions': successful_sessions,
            'total_resources': total_resources,
            'success_rate': round((successful_sessions / total_sessions * 100) if total_sessions > 0 else 0, 1),
        })
        
        return context


class KRTMakerView(FormView):
    """Main KRT maker form and processing view"""
    template_name = 'web/krt_maker.html'
    form_class = KRTMakerForm
    
    def get_context_data(self, **kwargs):
        """Add model choices to context for JavaScript synchronization"""
        context = super().get_context_data(**kwargs)
        
        # Add model choices for JavaScript - ensures Django and JS stay synchronized
        from .forms import KRTMakerForm
        form = KRTMakerForm()
        context['model_choices'] = {
            'anthropic': form.ANTHROPIC_MODELS,
            'gemini': form.GEMINI_MODELS,
            'openai_compatible': form.OPENAI_COMPATIBLE_MODELS,
        }
        
        return context
    
    def form_valid(self, form):
        """Process the form and extract KRT"""
        # Generate unique session ID
        session_id = uuid.uuid4().hex[:16]
        
        # Get form data
        input_method = form.cleaned_data['input_method']
        xml_file = form.cleaned_data.get('xml_file')
        biorxiv_url = form.cleaned_data.get('biorxiv_url')
        mode = form.cleaned_data['mode']
        provider = form.cleaned_data.get('provider')
        model = form.cleaned_data.get('model')
        base_url = form.cleaned_data.get('base_url')
        api_key = form.cleaned_data.get('api_key')
        extra_instructions = form.cleaned_data.get('extra_instructions')
        
        # Handle input method - get XML file path and metadata
        xml_path = None
        original_filename = "unknown.xml"
        file_size = 0
        temp_xml_path = None  # For cleanup if downloaded
        temp_path = None  # For Django file storage cleanup (uploads only)
        session = None  # Initialize session variable for exception handling
        local_xml_file = None  # Initialize local_xml_file for local storage reference
        
        try:
            if input_method == 'upload' and xml_file:
                # File upload method
                original_filename = xml_file.name
                file_size = xml_file.size
                
                # Save uploaded file to system temp directory (NOT in project media directory)
                # This prevents Django auto-reloader from detecting the file and restarting
                temp_file = tempfile.NamedTemporaryFile(
                    mode='wb',
                    suffix=f"_{original_filename}",
                    delete=False,
                    dir=tempfile.gettempdir()  # Use system temp directory
                )
                
                # Copy uploaded file content to temp file
                for chunk in xml_file.chunks():
                    temp_file.write(chunk)
                temp_file.close()
                
                xml_path = temp_file.name
                temp_xml_path = xml_path  # For cleanup via finally block
                # Note: For uploads, we don't use Django's default_storage since 
                # we're saving to system temp directory to avoid auto-reload issues
                
            elif input_method == 'url' and biorxiv_url:
                # bioRxiv URL method
                fetcher = BioRxivFetcher()
                
                # Parse DOI from URL
                doi = fetcher.parse_biorxiv_identifier(biorxiv_url)
                if not doi:
                    raise ValueError("Invalid bioRxiv URL or DOI")
                
                # Step 1: Check if XML file exists locally
                local_xml_file = XMLFile.objects.filter(doi=doi, is_available=True).first()
                
                if not local_xml_file:
                    raise ValueError(f"XML file not available in local dataset for {doi}. This paper may not be in our downloaded collection. Please run the XML download command to include more papers, or use the file upload method instead.")
                
                # Step 2: Verify file exists on disk
                if not local_xml_file.verify_file_exists():
                    raise ValueError(f"XML file for {doi} was found in database but is missing from disk. Please run the XML download command to re-download this file.")
                
                # Step 3: Use local XML file
                xml_path = local_xml_file.full_file_path
                temp_xml_path = None  # No cleanup needed for local files
                original_filename = f"{doi.replace('10.1101/', '')}.xml"
                file_size = local_xml_file.file_size
                
                print(f"✅ Using local XML file for {doi}")
                print(f"📄 File size: {file_size} bytes")
                print(f"📁 File path: {xml_path}")
                
            else:
                raise ValueError("No valid input method provided")
            
            # Get bioRxiv metadata if available
            biorxiv_metadata = {}
            session_doi = None
            if input_method == 'url' and biorxiv_url:
                fetcher = BioRxivFetcher()
                session_doi = fetcher.parse_biorxiv_identifier(biorxiv_url)
                if session_doi and local_xml_file:
                    # Use metadata from local XML file record
                    biorxiv_metadata = {
                        'doi': local_xml_file.doi,
                        'title': local_xml_file.title,
                        'authors': local_xml_file.authors,  # This is already JSON string
                        'authors_detailed': json.loads(local_xml_file.authors) if local_xml_file.authors else [],
                        'publication_date': local_xml_file.publication_date.strftime('%Y-%m-%d') if local_xml_file.publication_date else None,
                        'abstract': None,  # Abstract not stored in XMLFile model - will be extracted from XML
                        'journal': local_xml_file.journal,
                        'keywords': None,  # Keywords not stored in XMLFile model
                        'pmcid': None,
                        'pmid': None,
                        'source': 'local_xml_storage'
                    }
            
            # Parse authors - handle both list and string formats
            authors_list = []
            if biorxiv_metadata.get('authors_detailed'):
                # Use detailed author list if available
                authors_list = biorxiv_metadata['authors_detailed']
            elif biorxiv_metadata.get('authors'):
                # Handle both list and comma-separated string formats
                authors_data = biorxiv_metadata['authors']
                if isinstance(authors_data, list):
                    authors_list = authors_data
                elif isinstance(authors_data, str):
                    authors_list = [author.strip() for author in authors_data.split(',')]
                
            # Parse publication date if available  
            pub_date = None
            if biorxiv_metadata.get('publication_date'):
                try:
                    from datetime import datetime
                    date_str = biorxiv_metadata['publication_date']
                    # Europe PMC returns dates in YYYY-MM-DD format
                    pub_date = datetime.strptime(date_str[:10], '%Y-%m-%d').date()
                except (ValueError, TypeError):
                    pub_date = None
            
            # Create session record
            session = KRTSession.objects.create(
                session_id=session_id,
                original_filename=original_filename,
                file_size=file_size,
                input_method=input_method,
                doi=biorxiv_metadata.get('doi') or session_doi,
                biorxiv_id=session_doi,  # Store the bioRxiv ID
                epmc_id=None,  # Will be populated when we get EPMC ID
                authors=json.dumps(authors_list) if authors_list else None,
                publication_date=pub_date,
                journal=biorxiv_metadata.get('journal', 'bioRxiv'),
                keywords=json.dumps(biorxiv_metadata.get('keywords', [])) if biorxiv_metadata.get('keywords') else None,
                categories=None,  # Categories not available from Europe PMC
                mode=mode,
                provider=provider,
                model_name=model,
                base_url=base_url,
                extra_instructions=extra_instructions,
                status='processing'
            )
            
            # Note: No ProcessedFile record needed since we use system temp files
            # This avoids Django auto-reload issues from file monitoring
            
            # Validate XML file
            validate_xml_file(xml_path)
            
            # Detect existing KRT in the article (import only when needed to avoid auto-reload)
            try:
                from krt_detector import detect_existing_krt, format_krt_data_for_display
                existing_krt_info = detect_existing_krt(xml_path)
                existing_krt_data = format_krt_data_for_display(existing_krt_info.get('krt_tables', []))
                
                # Update session with existing KRT info
                session.existing_krt_detected = existing_krt_info.get('has_krt', False)
                session.existing_krt_count = existing_krt_info.get('krt_count', 0)
                session.existing_krt_data = existing_krt_data
            except Exception as e:
                print(f"KRT detection failed: {e}")
                # Set default values if KRT detection fails
                session.existing_krt_detected = False
                session.existing_krt_count = 0
                session.existing_krt_data = []
            
            session.save()
            
            # Build options for KRT extraction
            options = BuildOptions(
                mode='llm' if mode == 'llm' else 'regex',
                provider=provider,
                model=model,
                base_url=base_url,
                api_key=api_key,
                extra_instructions=extra_instructions,
            )
            
            # Validate LLM configuration if needed
            if mode == 'llm':
                validate_api_config(provider or "openai", api_key, model)
            
            # Record start time
            start_time = time.time()
            
            # Extract KRT
            result = build_from_xml_path(xml_path, options)
            
            # Record processing time
            processing_time = time.time() - start_time
            
            # Update session with results
            session.status = 'completed'
            # Use bioRxiv metadata title if available, otherwise use extracted title
            session.title = biorxiv_metadata.get('title') or result.get('title', '')
            # Use bioRxiv metadata abstract if available, otherwise use extracted abstract
            session.abstract = biorxiv_metadata.get('abstract') or result.get('abstract', '')
            session.krt_data = result.get('rows', [])
            session.processing_time = processing_time
            session.save()
            
            # Update analytics
            session.update_analytics()
            
            # Note: Temp files cleaned up in finally block below
            
            messages.success(
                self.request, 
                f'KRT extraction completed successfully! Found {len(result.get("rows", []))} resources in {processing_time:.1f} seconds.'
            )
            
            # Redirect to article profile instead of results page
            if session.doi:
                return redirect('web:article_profile', identifier=session.doi)
            else:
                return redirect('web:article_profile', identifier=session_id)
            
        except KRTValidationError as e:
            if session:
                session.status = 'failed'
                session.error_message = str(e)
                session.save()
            
            messages.error(self.request, f'Validation error: {e}')
            return self.form_invalid(form)
            
        except Exception as e:
            if session:
                session.status = 'failed'
                session.error_message = str(e)
                session.save()
            
            messages.error(self.request, f'Processing error: {e}')
            return self.form_invalid(form)
            
        finally:
            # Clean up temporary bioRxiv XML file
            if temp_xml_path and os.path.exists(temp_xml_path):
                try:
                    os.unlink(temp_xml_path)
                except Exception:
                    pass  # Ignore cleanup errors


class ResultsView(TemplateView):
    """Display KRT extraction results"""
    template_name = 'web/results.html'
    
    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        session_id = kwargs.get('session_id')
        
        try:
            session = get_object_or_404(KRTSession, session_id=session_id)
            context['session'] = session
            
            if session.status == 'completed':
                # Process KRT data for display
                krt_data = session.krt_data or []
                
                # Group by resource type for better visualization
                resource_types = {}
                for row in krt_data:
                    resource_type = row.get('RESOURCE TYPE', 'Other')
                    if resource_type not in resource_types:
                        resource_types[resource_type] = []
                    resource_types[resource_type].append(row)
                
                # Add KRT validation
                validation_warnings = validate_krt_completeness(krt_data)
                quality_score, max_score, quality_notes = get_krt_quality_score(krt_data)
                improvement_suggestions = suggest_krt_improvements(krt_data)
                
                context.update({
                    'krt_data': krt_data,
                    'resource_types': resource_types,
                    'resource_count': len(krt_data),
                    'new_count': len([r for r in krt_data if r.get('NEW/REUSE', '').lower() == 'new']),
                    'reuse_count': len([r for r in krt_data if r.get('NEW/REUSE', '').lower() == 'reuse']),
                    'validation_warnings': validation_warnings,
                    'quality_score': quality_score,
                    'max_quality_score': max_score,
                    'quality_notes': quality_notes,
                    'improvement_suggestions': improvement_suggestions,
                    'quality_percentage': int((quality_score / max_score) * 100) if max_score > 0 else 0,
                })
            
        except KRTSession.DoesNotExist:
            messages.error(self.request, 'Session not found or expired.')
            return redirect('web:home')
        
        return context


class ArticleDashboardView(TemplateView):
    """Dashboard showing all processed articles with their metadata and LLM comparison"""
    template_name = 'web/article_dashboard.html'
    
    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        
        # Get unique articles using the class method
        unique_articles = KRTSession.get_unique_articles()
        
        # Statistics
        total_articles = len(unique_articles)
        total_sessions = KRTSession.objects.filter(status='completed').count()
        biorxiv_articles = len([a for a in unique_articles if a['primary_session'].is_biorxiv_paper])
        
        # LLM usage statistics
        llm_stats = {}
        for article in unique_articles:
            for llm_result in article['llm_results']:
                provider_model = f"{llm_result['provider']}_{llm_result['model']}"
                if provider_model not in llm_stats:
                    llm_stats[provider_model] = {
                        'count': 0, 
                        'avg_resources': 0, 
                        'avg_time': 0,
                        'provider': llm_result['provider'],
                        'model': llm_result['model']
                    }
                llm_stats[provider_model]['count'] += 1
                llm_stats[provider_model]['avg_resources'] += llm_result['resources_found']
                llm_stats[provider_model]['avg_time'] += llm_result['processing_time'] or 0
        
        # Calculate averages
        for stats in llm_stats.values():
            if stats['count'] > 0:
                stats['avg_resources'] = round(stats['avg_resources'] / stats['count'], 1)
                stats['avg_time'] = round(stats['avg_time'] / stats['count'], 2)
        
        context.update({
            'articles': unique_articles,
            'total_articles': total_articles,
            'total_sessions': total_sessions,
            'biorxiv_articles': biorxiv_articles,
            'upload_articles': total_articles - biorxiv_articles,
            'llm_stats': list(llm_stats.values()),
        })
        
        return context


class ArticleProfileView(TemplateView):
    """Detailed profile page for a specific article showing all LLM results and metadata"""
    template_name = 'web/article_profile.html'
    
    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        
        # Get article identifier from URL
        identifier = kwargs.get('identifier')
        
        # Find sessions for this article (by DOI or session_id)
        if identifier.startswith('10.1101/') or '/' in identifier:
            # DOI identifier (bioRxiv DOIs start with 10.1101/)
            sessions = KRTSession.objects.filter(doi=identifier, status='completed').order_by('-created_at')
            article_key = identifier
        else:
            # Session ID identifier - find by session and then group by DOI/filename
            try:
                primary_session = KRTSession.objects.get(session_id=identifier, status='completed')
                article_key = primary_session.doi or primary_session.original_filename
                sessions = KRTSession.objects.filter(
                    Q(doi=article_key) | Q(original_filename=article_key),
                    status='completed'
                ).order_by('-created_at')
            except KRTSession.DoesNotExist:
                messages.error(self.request, 'Article not found.')
                return redirect('web:article_dashboard')
        
        if not sessions.exists():
            messages.error(self.request, 'Article not found.')
            return redirect('web:article_dashboard')
        
        # Primary session (most recent or best)
        primary_session = sessions.first()
        
        # Group sessions by LLM model
        llm_results = {}
        regex_results = []
        
        for session in sessions:
            if session.mode == 'llm':
                key = f"{session.provider}_{session.model_name}"
                if key not in llm_results:
                    llm_results[key] = {
                        'provider': session.provider,
                        'model': session.model_name,
                        'sessions': [],
                        'best_session': session,
                        'avg_resources': 0,
                        'avg_time': 0
                    }
                
                llm_results[key]['sessions'].append(session)
                
                # Update best session
                if session.resources_found > llm_results[key]['best_session'].resources_found:
                    llm_results[key]['best_session'] = session
            else:
                regex_results.append(session)
        
        # Calculate averages for each LLM
        for result in llm_results.values():
            sessions_list = result['sessions']
            result['avg_resources'] = sum(s.resources_found for s in sessions_list) / len(sessions_list)
            result['avg_time'] = sum(s.processing_time or 0 for s in sessions_list) / len(sessions_list)
        
        # Check for existing KRT in the article
        existing_krt_info = self._detect_existing_krt(primary_session)
        
        # Get admin-generated KRTs for this article/DOI
        admin_krts = []
        if primary_session.doi:
            admin_krts = AdminKRT.get_for_doi(primary_session.doi)
        
        # Get best admin KRT for quick comparison
        best_admin_krt = None
        if primary_session.doi:
            best_admin_krt = AdminKRT.get_best_for_doi(primary_session.doi)
        
        context.update({
            'primary_session': primary_session,
            'sessions': sessions,
            'all_sessions': sessions,
            'llm_results': list(llm_results.values()),
            'regex_results': regex_results,
            'total_sessions': sessions.count(),
            'article_key': article_key,
            'existing_krt_info': existing_krt_info,
            'admin_krts': admin_krts,
            'best_admin_krt': best_admin_krt,
        })
        
        return context
    
    def _detect_existing_krt(self, session):
        """Detect if the article already contains KRT tables"""
        return {
            'has_krt': session.existing_krt_detected,
            'krt_count': session.existing_krt_count,
            'krt_data': session.existing_krt_data,
            'confidence_score': 85  # Default confidence score, can be enhanced later
        }


class AdminKRTManagementView(TemplateView):
    """Admin interface for managing admin-generated KRTs"""
    template_name = 'web/admin_krt_management.html'
    
    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        
        # Get statistics about admin KRTs
        stats = AdminKRT.get_statistics()
        
        # Get recent admin KRTs
        recent_krts = AdminKRT.objects.select_related('xml_file').order_by('-created_at')[:20]
        
        # Get pending KRTs that need generation
        pending_krts = AdminKRT.get_pending_generation(limit=50)
        
        # Get featured/high-quality KRTs
        featured_krts = AdminKRT.objects.filter(
            is_featured=True, 
            is_public=True
        ).select_related('xml_file').order_by('-created_at')[:10]
        
        # Get provider/model breakdown
        provider_stats = {}
        for krt in AdminKRT.objects.values('provider', 'model_name').annotate(
            count=Count('id'),
            avg_resources=Avg('resources_found'),
            avg_time=Avg('processing_time')
        ):
            key = f"{krt['provider']}_{krt['model_name']}"
            provider_stats[key] = {
                'provider': krt['provider'],
                'model': krt['model_name'],
                'count': krt['count'],
                'avg_resources': round(krt['avg_resources'] or 0, 1),
                'avg_time': round(krt['avg_time'] or 0, 2)
            }
        
        context.update({
            'stats': stats,
            'recent_krts': recent_krts,
            'pending_krts': pending_krts,
            'featured_krts': featured_krts,
            'provider_stats': list(provider_stats.values()),
            'xml_file_count': XMLFile.objects.filter(is_available=True).count(),
        })
        
        return context


class DatabaseManagementView(TemplateView):
    """Database management interface for populating and managing the article database"""
    template_name = 'web/database_management.html'
    
    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        
        # Get database statistics
        total_articles = Article.objects.count()
        articles_with_sessions = Article.objects.filter(total_sessions__gt=0).count()
        articles_with_krt = Article.objects.filter(has_existing_krt=True).count()
        
        # Get latest articles
        recent_articles = Article.objects.order_by('-last_processed')[:10]
        
        # Get year statistics from existing articles
        year_distribution = {}
        for article in Article.objects.filter(publication_date__isnull=False):
            year = article.publication_date.year
            year_distribution[year] = year_distribution.get(year, 0) + 1
        
        # Get journal distribution
        journal_distribution = {}
        for article in Article.objects.filter(journal__isnull=False).exclude(journal=''):
            journal = article.journal
            journal_distribution[journal] = journal_distribution.get(journal, 0) + 1
        
        # Sort distributions
        year_distribution = dict(sorted(year_distribution.items(), reverse=True)[:10])
        journal_distribution = dict(sorted(journal_distribution.items(), key=lambda x: x[1], reverse=True)[:10])
        
        # Population status
        expected_total = 19405  # Based on our Europe PMC statistics
        population_percentage = (total_articles / expected_total * 100) if expected_total > 0 else 0
        
        context.update({
            'total_articles': total_articles,
            'articles_with_sessions': articles_with_sessions,
            'articles_with_krt': articles_with_krt,
            'articles_without_sessions': total_articles - articles_with_sessions,
            'recent_articles': recent_articles,
            'year_distribution': year_distribution,
            'journal_distribution': journal_distribution,
            'expected_total': expected_total,
            'population_percentage': population_percentage,
            'is_populated': total_articles > 1000,  # Consider populated if more than 1000 articles
        })
        
        return context
    
    def post(self, request, *args, **kwargs):
        """Handle database population requests"""
        action = request.POST.get('action')
        
        if action == 'populate_full':
            # Trigger full population (this would typically be done via background task)
            messages.info(request, '🚀 Full database population initiated. This may take several minutes. Check the admin interface for progress.')
            
            # In a production environment, you would typically queue this as a background task
            # For now, we'll provide instructions for manual execution
            messages.warning(request, '💡 For best performance, run this command manually: python manage.py populate_articles')
            
        elif action == 'populate_update':
            # Trigger update-only population
            messages.info(request, '🔄 Database update initiated. Only new articles will be added.')
            messages.warning(request, '💡 Run this command manually: python manage.py populate_articles --update-only')
            
        elif action == 'populate_year':
            year = request.POST.get('year')
            if year:
                messages.info(request, f'📅 Population for year {year} initiated.')
                messages.warning(request, f'💡 Run this command manually: python manage.py populate_articles --start-year {year} --end-year {year}')
        
        return redirect('web:database_management')


@require_http_methods(["GET"])
def export_krt(request, session_id, format_type):
    """Export KRT data in various formats"""
    try:
        session = KRTSession.objects.get(session_id=session_id, status='completed')
    except KRTSession.DoesNotExist:
        return JsonResponse({'error': 'Session not found'}, status=404)
    
    # Get KRT data
    krt_data = session.krt_data or []
    
    if not krt_data:
        return JsonResponse({'error': 'No KRT data available'}, status=404)
    
    # Record export for analytics
    KRTExport.objects.create(
        session=session,
        format=format_type,
        ip_address=request.META.get('REMOTE_ADDR'),
        user_agent=request.META.get('HTTP_USER_AGENT', '')[:500]
    )
    
    if format_type == 'json':
        response = JsonResponse({
            'title': session.display_title,
            'session_id': session_id,
            'doi': session.doi,
            'processing_date': session.created_at.isoformat(),
            'mode': session.mode,
            'provider': session.provider,
            'model': session.model_name,
            'resources_found': len(krt_data),
            'krt_data': krt_data
        }, json_dumps_params={'indent': 2})
        response['Content-Disposition'] = f'attachment; filename="krt_{session_id}.json"'
        return response
    
    elif format_type == 'csv':
        import csv
        from io import StringIO
        
        output = StringIO()
        if krt_data:
            # Use the keys from the first row, but ensure proper order
            fieldnames = [
                'RESOURCE TYPE', 'RESOURCE NAME', 'SOURCE', 
                'IDENTIFIER', 'NEW/REUSE', 'ADDITIONAL INFORMATION'
            ]
            writer = csv.DictWriter(output, fieldnames=fieldnames)
            writer.writeheader()
            
            for row in krt_data:
                # Create a clean row with all required fields
                clean_row = {}
                for field in fieldnames:
                    clean_row[field] = row.get(field, '')
                writer.writerow(clean_row)
        
        response = HttpResponse(output.getvalue(), content_type='text/csv')
        response['Content-Disposition'] = f'attachment; filename="krt_{session_id}.csv"'
        return response
    
    elif format_type == 'excel':
        try:
            import openpyxl
            from openpyxl.styles import Font, PatternFill
            from io import BytesIO
        except ImportError:
            return JsonResponse({'error': 'Excel export not available (openpyxl not installed)'}, status=500)
        
        # Create workbook
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Key Resources Table"
        
        # Add metadata
        ws['A1'] = f"Key Resources Table - {session.display_title}"
        ws['A1'].font = Font(bold=True, size=14)
        ws['A2'] = f"Session ID: {session_id}"
        ws['A3'] = f"DOI: {session.doi or 'N/A'}"
        ws['A4'] = f"Processed: {session.created_at.strftime('%Y-%m-%d %H:%M')}"
        ws['A5'] = f"Mode: {session.mode.upper()}"
        if session.provider:
            ws['A6'] = f"Provider: {session.provider} ({session.model_name})"
        
        # Add headers
        headers = [
            'RESOURCE TYPE', 'RESOURCE NAME', 'SOURCE', 
            'IDENTIFIER', 'NEW/REUSE', 'ADDITIONAL INFORMATION'
        ]
        header_row = 8
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=header_row, column=col, value=header)
            cell.font = Font(bold=True)
            cell.fill = PatternFill(start_color="CCCCCC", end_color="CCCCCC", fill_type="solid")
        
        # Add data
        for row_idx, data_row in enumerate(krt_data, header_row + 1):
            for col_idx, header in enumerate(headers, 1):
                ws.cell(row=row_idx, column=col_idx, value=data_row.get(header, ''))
        
        # Auto-adjust column widths
        for column in ws.columns:
            max_length = 0
            column_letter = column[0].column_letter
            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            adjusted_width = min(max_length + 2, 50)  # Cap at 50 characters
            ws.column_dimensions[column_letter].width = adjusted_width
        
        # Save to BytesIO
        output = BytesIO()
        wb.save(output)
        output.seek(0)
        
        response = HttpResponse(
            output.getvalue(),
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )
        response['Content-Disposition'] = f'attachment; filename="krt_{session_id}.xlsx"'
        return response
    
    else:
        return JsonResponse({'error': 'Invalid format type'}, status=400)


@require_http_methods(["POST"])
def check_doi_availability(request):
    """
    AJAX endpoint to check if a DOI has full text XML available for KRT extraction.
    Returns availability status and detailed information.
    """
    try:
        data = json.loads(request.body)
        doi_input = data.get('doi', '').strip()
        
        if not doi_input:
            return JsonResponse({
                'success': False,
                'error': 'No DOI provided'
            }, status=400)
        
        # Initialize fetcher
        fetcher = BioRxivFetcher()
        
        # Parse DOI from input (handles URLs and DOIs)
        doi = fetcher.parse_biorxiv_identifier(doi_input)
        if not doi:
            return JsonResponse({
                'success': False,
                'error': 'Invalid bioRxiv URL or DOI format',
                'details': 'Please provide a valid bioRxiv DOI (e.g., 10.1101/2025.01.01.123456) or URL'
            }, status=400)
        
        # Step 1: Search for EPMC ID
        epmc_id = fetcher.search_epmc_for_doi(doi)
        
        if not epmc_id:
            return JsonResponse({
                'success': False,
                'error': 'Paper not found in Europe PMC',
                'details': f'DOI {doi} was not found in the Europe PMC database. This could mean the paper is very new or not indexed yet.',
                'doi': doi,
                'indexed': False
            }, status=404)
        
        # Step 2: Check full text availability
        full_text_status = fetcher.check_full_text_availability(epmc_id)
        
        # Step 3: Get basic metadata for display
        metadata = fetcher.get_paper_metadata(doi) or {}
        
        # Prepare response
        response_data = {
            'success': True,
            'doi': doi,
            'epmc_id': epmc_id,
            'indexed': True,
            'full_text_available': full_text_status['available'],
            'can_extract_krt': full_text_status['available'],  # Only if full text is available
            'metadata': {
                'title': metadata.get('preprint_title', 'Title not available'),
                'authors': metadata.get('preprint_authors', 'Authors not available'),
                'date': metadata.get('preprint_date', 'Date not available'),
                'journal': metadata.get('preprint_platform', 'bioRxiv'),
                'abstract': metadata.get('preprint_abstract', '')[:200] + '...' if metadata.get('preprint_abstract') else 'Abstract not available'
            },
            'availability_details': full_text_status
        }
        
        # Add specific messages based on availability status
        if full_text_status['available']:
            response_data['message'] = f"✅ Full text XML is available for KRT extraction!"
            if full_text_status.get('content_length'):
                response_data['message'] += f" (File size: ~{full_text_status['content_length']} bytes)"
        else:
            if full_text_status.get('status_code') == 404:
                response_data['message'] = f"⚠️ Paper is indexed but full text XML is not yet available. This often happens with very recent preprints."
                response_data['suggestion'] = "Try again in a few days as Europe PMC processes new papers regularly."
            else:
                response_data['message'] = f"❌ Full text XML is not available. Error: {full_text_status.get('error', 'Unknown error')}"
                response_data['suggestion'] = "You may need to contact the authors directly for the manuscript."
        
        return JsonResponse(response_data)
        
    except json.JSONDecodeError:
        return JsonResponse({
            'success': False,
            'error': 'Invalid JSON in request body'
        }, status=400)
        
    except Exception as e:
        return JsonResponse({
            'success': False,
            'error': 'Server error while checking DOI availability',
            'details': str(e)
        }, status=500)


@require_http_methods(["GET"])
def get_krt_data_api(request, session_id):
    """
    API endpoint to retrieve KRT data by session ID for AJAX modal display.
    """
    try:
        # Get the session
        session = get_object_or_404(KRTSession, session_id=session_id)
        
        # Parse KRT data (handle both list and JSON string formats)
        krt_data = []
        if session.krt_data:
            if isinstance(session.krt_data, list):
                krt_data = session.krt_data
            elif isinstance(session.krt_data, str):
                try:
                    krt_data = json.loads(session.krt_data)
                except json.JSONDecodeError:
                    return JsonResponse({
                        'success': False,
                        'error': 'Invalid KRT data format'
                    }, status=500)
            else:
                krt_data = []
        
        # Calculate summary statistics
        summary = {
            'total_resources': len(krt_data),
            'new_resources': sum(1 for item in krt_data if item.get('NEW/REUSE', '').lower() == 'new'),
            'reused_resources': sum(1 for item in krt_data if item.get('NEW/REUSE', '').lower() == 'reuse'),
            'processing_time': float(session.processing_time) if session.processing_time else 0,
            'provider': session.provider,
            'model': session.model_name,
            'mode': session.mode,
            'created_at': session.created_at.isoformat(),
        }
        
        return JsonResponse({
            'success': True,
            'session_id': session_id,
            'krt_data': krt_data,
            'summary': summary,
            'session_info': {
                'title': session.display_title,
                'doi': session.doi,
                'is_biorxiv_paper': session.is_biorxiv_paper,
                'status': session.status,
            }
        })
        
    except KRTSession.DoesNotExist:
        return JsonResponse({
            'success': False,
            'error': f'Session not found: {session_id}'
        }, status=404)
    except Exception as e:
        return JsonResponse({
            'success': False,
            'error': f'Server error: {str(e)}'
        }, status=500)


@require_http_methods(["GET"])
def get_krt_data_by_doi_api(request, doi):
    """
    API endpoint to retrieve KRT data by DOI for external API access.
    Returns the best KRT result for the given DOI.
    """
    try:
        # Find the best session for this DOI (highest resource count)
        sessions = KRTSession.objects.filter(
            doi=doi,
            status='completed'
        ).exclude(
            krt_data__isnull=True
        ).exclude(
            krt_data=''
        ).order_by('-resources_found', '-created_at')
        
        if not sessions.exists():
            return JsonResponse({
                'success': False,
                'error': f'No completed KRT sessions found for DOI: {doi}'
            }, status=404)
        
        best_session = sessions.first()
        
        # Parse KRT data (handle both list and JSON string formats)
        krt_data = []
        if best_session.krt_data:
            if isinstance(best_session.krt_data, list):
                krt_data = best_session.krt_data
            elif isinstance(best_session.krt_data, str):
                try:
                    krt_data = json.loads(best_session.krt_data)
                except json.JSONDecodeError:
                    return JsonResponse({
                        'success': False,
                        'error': 'Invalid KRT data format'
                    }, status=500)
            else:
                krt_data = []
        
        # Calculate summary statistics
        summary = {
            'total_resources': len(krt_data),
            'new_resources': sum(1 for item in krt_data if item.get('NEW/REUSE', '').lower() == 'new'),
            'reused_resources': sum(1 for item in krt_data if item.get('NEW/REUSE', '').lower() == 'reuse'),
            'processing_time': float(best_session.processing_time) if best_session.processing_time else 0,
            'provider': best_session.provider,
            'model': best_session.model_name,
            'mode': best_session.mode,
            'created_at': best_session.created_at.isoformat(),
            'total_sessions': sessions.count(),
        }
        
        return JsonResponse({
            'success': True,
            'doi': doi,
            'session_id': best_session.session_id,
            'krt_data': krt_data,
            'summary': summary,
            'article_info': {
                'title': best_session.display_title,
                'authors': best_session.formatted_authors,
                'publication_date': best_session.publication_date.isoformat() if best_session.publication_date else None,
                'journal': best_session.journal,
                'is_biorxiv_paper': best_session.is_biorxiv_paper,
            }
        })
        
    except Exception as e:
        return JsonResponse({
            'success': False,
            'error': f'Server error: {str(e)}'
        }, status=500)


@require_http_methods(["GET"])
def get_doi_suggestions(request):
    """
    AJAX endpoint to get DOI suggestions organized by year.
    Returns random preprints from each available year with load more functionality.
    Load more batch sizes: 25 -> 100 -> 200
    """
    try:
        # Get parameters
        query = request.GET.get('q', '').strip()
        load_more = request.GET.get('load_more', '').lower() == 'true'
        load_count = int(request.GET.get('load_count', '0'))
        specific_year = request.GET.get('year', '').strip()
        
        # Determine batch size based on load count
        # First load: 10 per year for overview, then 25, 100, 200 for load more
        if load_more:
            if load_count == 0:
                batch_size = 25
            elif load_count == 1:
                batch_size = 100
            elif load_count >= 2:
                batch_size = 200
            else:
                batch_size = 25
        else:
            batch_size = 10  # Initial load: 10 per year for overview
        
        # Base queryset of available XML files
        xml_files = XMLFile.objects.filter(is_available=True)
        
        if query:
            # If user is typing, filter by DOI or title containing the query
            limit = min(batch_size, 50)  # Cap search results at 50
            xml_files = xml_files.filter(
                Q(doi__icontains=query) | 
                Q(title__icontains=query)
            ).order_by('doi')[:limit]
            
            suggestions = []
            for xml_file in xml_files:
                suggestions.append(_format_xml_file_for_suggestion(xml_file))
            
            return JsonResponse({
                'success': True,
                'suggestions': suggestions,
                'total_available': XMLFile.objects.filter(is_available=True).count(),
                'query': query,
                'by_year': False,
                'load_more': False
            })
        
        elif specific_year:
            # Load more for a specific year
            if specific_year == 'unknown':
                year_files = xml_files.filter(publication_date__isnull=True)
            else:
                try:
                    year_int = int(specific_year)
                    year_files = xml_files.filter(publication_date__year=year_int)
                except ValueError:
                    return JsonResponse({
                        'success': False,
                        'error': 'Invalid year parameter'
                    }, status=400)
            
            # Get random papers for this year
            random_files = year_files.order_by('?')[:batch_size]
            
            suggestions = []
            for xml_file in random_files:
                suggestions.append(_format_xml_file_for_suggestion(xml_file))
            
            total_in_year = year_files.count()
            can_load_more = total_in_year > len(suggestions) and load_count < 3
            
            return JsonResponse({
                'success': True,
                'suggestions': suggestions,
                'year': specific_year,
                'total_in_year': total_in_year,
                'batch_size': batch_size,
                'load_count': load_count,
                'can_load_more': can_load_more,
                'next_batch_size': 100 if load_count == 0 else 200 if load_count == 1 else None,
                'by_year': False,
                'load_more': True
            })
        
        else:
            # Initial load: overview with 10 random suggestions from each year
            suggestions_by_year = {}
            
            # Get available years (from publication_date)
            years = XMLFile.objects.filter(
                is_available=True,
                publication_date__isnull=False
            ).values_list(
                'publication_date__year', flat=True
            ).distinct().order_by('-publication_date__year')
            
            total_suggestions = 0
            for year in years:
                year_files = xml_files.filter(publication_date__year=year)
                year_count = year_files.count()
                
                # Get random papers for overview (10 per year)
                year_random = year_files.order_by('?')[:10]
                
                year_suggestions = []
                for xml_file in year_random:
                    year_suggestions.append(_format_xml_file_for_suggestion(xml_file))
                
                if year_suggestions:
                    suggestions_by_year[str(year)] = {
                        'year': year,
                        'total_count': year_count,
                        'suggestions': year_suggestions,
                        'can_load_more': year_count > 10,
                        'showing_count': len(year_suggestions)
                    }
                    total_suggestions += len(year_suggestions)
            
            # Also get papers with no publication date
            no_date_files = xml_files.filter(publication_date__isnull=True)
            no_date_count = no_date_files.count()
            
            if no_date_count > 0:
                no_date_random = no_date_files.order_by('?')[:10]
                no_date_suggestions = []
                for xml_file in no_date_random:
                    no_date_suggestions.append(_format_xml_file_for_suggestion(xml_file))
                
                suggestions_by_year['unknown'] = {
                    'year': 'Unknown',
                    'total_count': no_date_count,
                    'suggestions': no_date_suggestions,
                    'can_load_more': no_date_count > 10,
                    'showing_count': len(no_date_suggestions)
                }
                total_suggestions += len(no_date_suggestions)
            
            return JsonResponse({
                'success': True,
                'suggestions_by_year': suggestions_by_year,
                'total_suggestions': total_suggestions,
                'total_available': XMLFile.objects.filter(is_available=True).count(),
                'query': None,
                'by_year': True,
                'load_more': False
            })
        
    except Exception as e:
        return JsonResponse({
            'success': False,
            'error': 'Server error while getting DOI suggestions',
            'details': str(e)
        }, status=500)


@require_http_methods(["POST"])
def check_doi_local_availability(request):
    """
    AJAX endpoint to check if a DOI has XML file available locally.
    This replaces the Europe PMC availability check for local XML storage.
    """
    try:
        data = json.loads(request.body)
        doi_input = data.get('doi', '').strip()
        
        if not doi_input:
            return JsonResponse({
                'success': False,
                'error': 'No DOI provided'
            }, status=400)
        
        # Initialize fetcher for DOI parsing
        fetcher = BioRxivFetcher()
        
        # Parse DOI from input (handles URLs and DOIs)
        doi = fetcher.parse_biorxiv_identifier(doi_input)
        if not doi:
            return JsonResponse({
                'success': False,
                'error': 'Invalid bioRxiv URL or DOI format',
                'details': 'Please provide a valid bioRxiv DOI (e.g., 10.1101/2025.01.01.123456) or URL'
            }, status=400)
        
        # Check if XML file exists locally
        xml_file = XMLFile.objects.filter(doi=doi, is_available=True).first()
        
        if xml_file:
            # Verify file still exists on disk
            if xml_file.verify_file_exists():
                # Format authors for display
                authors_display = "Unknown Authors"
                if xml_file.authors:
                    try:
                        authors_list = json.loads(xml_file.authors)
                        if len(authors_list) > 3:
                            authors_display = f"{', '.join(authors_list[:3])} et al."
                        else:
                            authors_display = ', '.join(authors_list)
                    except:
                        authors_display = xml_file.authors[:50] + "..." if len(xml_file.authors) > 50 else xml_file.authors
                
                return JsonResponse({
                    'success': True,
                    'doi': doi,
                    'local_file_available': True,
                    'can_extract_krt': True,
                    'metadata': {
                        'title': xml_file.title or 'Unknown Title',
                        'authors': authors_display,
                        'date': xml_file.publication_date.strftime('%Y-%m-%d') if xml_file.publication_date else 'Unknown Date',
                        'journal': xml_file.journal,
                        'file_size': xml_file.file_size,
                        'downloaded_at': xml_file.downloaded_at.strftime('%Y-%m-%d %H:%M') if xml_file.downloaded_at else None
                    },
                    'message': f"✅ XML file available locally! Ready for KRT extraction.",
                    'source': 'local_storage'
                })
            else:
                return JsonResponse({
                    'success': True,
                    'doi': doi,
                    'local_file_available': False,
                    'can_extract_krt': False,
                    'message': f"❌ XML file was downloaded but is no longer available on disk.",
                    'suggestion': "Please run the XML download command to re-download this file.",
                    'source': 'local_storage'
                })
        else:
            return JsonResponse({
                'success': True,
                'doi': doi,
                'local_file_available': False,
                'can_extract_krt': False,
                'message': f"⚠️ XML file not available in local dataset.",
                'suggestion': "This paper may not be in our downloaded dataset. Please run the XML download command to include more papers.",
                'source': 'local_storage'
            })
        
    except json.JSONDecodeError:
        return JsonResponse({
            'success': False,
            'error': 'Invalid JSON in request body'
        }, status=400)
        
    except Exception as e:
        return JsonResponse({
            'success': False,
            'error': 'Server error while checking local DOI availability',
            'details': str(e)
        }, status=500)


def _format_xml_file_for_suggestion(xml_file):
    """Helper function to format XMLFile object for suggestion response"""
    # Format authors for display
    authors_display = "Unknown Authors"
    if xml_file.authors:
        try:
            authors_list = json.loads(xml_file.authors)
            if len(authors_list) > 3:
                authors_display = f"{', '.join(authors_list[:3])} et al."
            else:
                authors_display = ', '.join(authors_list)
        except:
            authors_display = xml_file.authors[:50] + "..." if len(xml_file.authors) > 50 else xml_file.authors
    
    return {
        'doi': xml_file.doi,
        'title': xml_file.title[:80] + "..." if xml_file.title and len(xml_file.title) > 80 else xml_file.title or "Unknown Title",
        'authors': authors_display,
        'publication_date': xml_file.publication_date.strftime('%Y-%m-%d') if xml_file.publication_date else 'Unknown Date',
        'file_size': xml_file.file_size,
        'downloaded_at': xml_file.downloaded_at.strftime('%Y-%m-%d') if xml_file.downloaded_at else None
    }


class AboutView(TemplateView):
    """About page with information about KRT Maker"""
    template_name = 'web/about.html'
    
    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        
        # Add some statistics for the about page
        from .models import KRTSession, Article, XMLFile
        
        stats = {
            'total_papers_processed': KRTSession.objects.filter(status='completed').count(),
            'total_papers_available': XMLFile.objects.filter(is_available=True).count(),
            'total_articles': Article.objects.count(),
            'avg_processing_time': KRTSession.objects.filter(
                status='completed', 
                processing_time__isnull=False
            ).aggregate(avg_time=models.Avg('processing_time'))['avg_time'] or 0,
        }
        
        context['stats'] = stats
        return context


class APIDocsView(TemplateView):
    """API Documentation page"""
    template_name = 'web/api_docs.html'
    
    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        
        # Add API endpoint information
        context['base_url'] = self.request.build_absolute_uri('/api/')
        context['version'] = '1.0'
        
        return context
def browser_extension_download(request):
    """Browser Extension Download page"""
    return render(request, 'web/browser_extension.html')
